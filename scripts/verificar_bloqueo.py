"""
Verificador cross-platform (openpyxl) del bloqueo de columnas en el .xlsm generado.

Confirma, en cada hoja de parcial, los criterios de aceptación:
  1. El rango "Producto del Parcial" -> "Porcentaje de asistencia" (por encabezado)
     queda bloqueado (locked=True) en las filas de datos 9-53.
  2. Las columnas previamente bloqueadas (BD:BJ = 56-62) siguen bloqueadas.
  3. Las celdas de captura a la izquierda del rango siguen editables (locked=False).
  4. La protección de hoja está activa (ws.protection.sheet is True).

A diferencia del generador (win32com / Windows), este script usa openpyxl y corre
en cualquier sistema operativo, sobre un archivo ya generado.

Uso:
    python scripts/verificar_bloqueo.py <ruta_al_xlsm>

Nota (openpyxl): una celda con estilo por defecto reporta locked=True (default de
Excel); las celdas desbloqueadas explícitamente reportan locked=False. En celdas
combinadas la protección vive en la esquina superior izquierda.
"""

import sys
import unicodedata

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: falta 'openpyxl'. Instala con: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


FILA_ENCABEZADO = 6
FILA_DATOS_INI = 9
FILA_DATOS_FIN = 53  # inclusive
HEADER_INI = "Producto del Parcial"
HEADER_FIN = "Porcentaje de asistencia"
# Rango previo que ya estaba bloqueado (BD..BJ = 56..62).
PREV_COL_INI, PREV_COL_FIN = 56, 62
# Columnas de captura a muestrear (deben seguir editables): D, Z, AR.
MUESTRA_EDITABLES = [4, 26, 44]


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def buscar_columna_por_encabezado(ws, texto, fila_encabezado=FILA_ENCABEZADO, max_col=200):
    objetivo = _norm(texto)
    tokens = objetivo.split()
    aprox = None
    for col in range(1, max_col + 1):
        val = _norm(ws.cell(row=fila_encabezado, column=col).value)
        if not val:
            continue
        if val == objetivo:
            return col
        if aprox is None and all(t in val.split() for t in tokens):
            aprox = col
    return aprox


def es_locked(cell):
    """True si la celda está bloqueada (default de Excel es locked=True)."""
    prot = cell.protection
    return bool(prot.locked) if prot is not None and prot.locked is not None else True


def verificar_hoja(ws):
    ok = True
    print(f"\n=== Hoja: {ws.title} ===")

    # 4. Protección de hoja activa
    protegida = bool(ws.protection.sheet)
    print(f"  [{'OK' if protegida else 'FALLA'}] Protección de hoja activa: {protegida}")
    ok &= protegida

    # Localizar rango por encabezado
    col_ini = buscar_columna_por_encabezado(ws, HEADER_INI)
    col_fin = buscar_columna_por_encabezado(ws, HEADER_FIN)
    if not col_ini or not col_fin or col_ini > col_fin:
        print(f"  [FALLA] No se localizó el rango por encabezado "
              f"('{HEADER_INI}'={col_ini}, '{HEADER_FIN}'={col_fin}).")
        return False
    print(f"  Rango por encabezado: {get_column_letter(col_ini)}..{get_column_letter(col_fin)} "
          f"(cols {col_ini}-{col_fin})")

    # 1. Rango objetivo bloqueado en filas 9-53
    fallos = []
    for row in range(FILA_DATOS_INI, FILA_DATOS_FIN + 1):
        for col in range(col_ini, col_fin + 1):
            if not es_locked(ws.cell(row=row, column=col)):
                fallos.append(f"{get_column_letter(col)}{row}")
    if fallos:
        ok = False
        muestra = ", ".join(fallos[:10]) + (" ..." if len(fallos) > 10 else "")
        print(f"  [FALLA] Rango '{HEADER_INI}'->'{HEADER_FIN}' bloqueado: "
              f"{len(fallos)} celdas NO bloqueadas ({muestra})")
    else:
        print(f"  [OK] Rango '{HEADER_INI}'->'{HEADER_FIN}' (filas "
              f"{FILA_DATOS_INI}-{FILA_DATOS_FIN}) totalmente bloqueado")

    # 2. Rango previo (BD:BJ) sigue bloqueado
    fallos_prev = []
    for row in range(FILA_DATOS_INI, FILA_DATOS_FIN + 1):
        for col in range(PREV_COL_INI, PREV_COL_FIN + 1):
            if not es_locked(ws.cell(row=row, column=col)):
                fallos_prev.append(f"{get_column_letter(col)}{row}")
    if fallos_prev:
        ok = False
        print(f"  [FALLA] Columnas previas BD:BJ: {len(fallos_prev)} celdas NO bloqueadas")
    else:
        print(f"  [OK] Columnas previas BD:BJ (56-62) siguen bloqueadas")

    # 3. Celdas de captura a la izquierda del rango siguen editables
    editables_ok = True
    for col in MUESTRA_EDITABLES:
        if col >= col_ini:
            continue  # solo muestreamos a la izquierda del rango bloqueado
        ref = f"{get_column_letter(col)}{FILA_DATOS_INI}"
        if es_locked(ws.cell(row=FILA_DATOS_INI, column=col)):
            editables_ok = False
            print(f"  [FALLA] Celda de captura {ref} quedó bloqueada (debía ser editable)")
    if editables_ok:
        refs = ", ".join(f"{get_column_letter(c)}{FILA_DATOS_INI}"
                         for c in MUESTRA_EDITABLES if c < col_ini)
        print(f"  [OK] Celdas de captura siguen editables ({refs})")
    ok &= editables_ok

    return ok


def main():
    if len(sys.argv) != 2:
        print("Uso: python scripts/verificar_bloqueo.py <ruta_al_xlsm>", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    try:
        wb = load_workbook(path, keep_vba=True)
    except Exception as e:
        print(f"Error al abrir '{path}': {e}", file=sys.stderr)
        sys.exit(2)

    hojas_parcial = [ws for ws in wb.worksheets if ws.title.lower().startswith("parcial")]
    if not hojas_parcial:
        print("Error: no se encontraron hojas 'Parcial_*' en el archivo.", file=sys.stderr)
        sys.exit(2)

    todo_ok = True
    for ws in hojas_parcial:
        todo_ok &= verificar_hoja(ws)

    print("\n" + ("=" * 40))
    if todo_ok:
        print("RESULTADO: ✅ TODOS los criterios se cumplen.")
        sys.exit(0)
    else:
        print("RESULTADO: ❌ Hay criterios que NO se cumplen (ver arriba).")
        sys.exit(1)


if __name__ == "__main__":
    main()
